import csv
import io
from datetime import date
from typing import Optional
from fastapi import APIRouter, Depends, Response
from fastapi.responses import HTMLResponse
from sqlalchemy.orm import Session
from ..database import get_db
from ..models import Hotel, WeeklyPerformance, DailyBooking
from ..services.kpi_engine import get_portfolio_summary, get_hotel_kpis, get_portfolio_trend
from ..services.pdf_generator import build_pdf_html
from ..auth import require_auth

router = APIRouter(prefix="/reports", tags=["reports"])


def _filtered_hotels(db, market, property_type, status):
    q = db.query(Hotel)
    if market:        q = q.filter(Hotel.market == market)
    if property_type: q = q.filter(Hotel.property_type == property_type)
    if status:        q = q.filter(Hotel.pipeline_status == status)
    return q.all()


def _period_label(date_from, date_to) -> str:
    if not date_from and not date_to:
        return "Season to date"
    fmt = lambda d: d.strftime("%d %b %Y")  # noqa: E731
    if date_from and date_to:
        return f"{fmt(date_from)} – {fmt(date_to)}"
    return f"From {fmt(date_from)}" if date_from else f"Until {fmt(date_to)}"


@router.get("/pdf")
def generate_pdf(
    market: Optional[str] = None,
    property_type: Optional[str] = None,
    status: Optional[str] = None,
    date_from: Optional[date] = None,
    date_to: Optional[date] = None,
    db: Session = Depends(get_db),
    _: str = Depends(require_auth),
):
    """
    Full investor report as print-ready HTML (auto-opens the browser's
    Save-as-PDF dialog): cover page, portfolio summary with all top-level KPIs,
    GMV trend chart, hotel list table, and one fact page per active hotel.
    """
    summary = get_portfolio_summary(
        db, market=market, property_type=property_type, status=status,
        date_from=date_from, date_to=date_to,
    )
    hotels = _filtered_hotels(db, market, property_type, status)
    hotel_kpis = [get_hotel_kpis(db, h, date_from=date_from, date_to=date_to) for h in hotels]
    trend = get_portfolio_trend(
        db, market=market, property_type=property_type, status=status,
        date_from=date_from, date_to=date_to,
    )

    html = build_pdf_html(summary, hotel_kpis, trend=trend,
                          period_label=_period_label(date_from, date_to))
    return HTMLResponse(content=html)


# Full per-hotel KPI columns shared by CSV and the Excel "Hotels" sheet
_HOTEL_COLS = [
    ("hotel_name", lambda k: k.hotel_name),
    ("market", lambda k: k.market),
    ("property_type", lambda k: k.property_type),
    ("room_count", lambda k: k.room_count),
    ("pipeline_status", lambda k: k.pipeline_status),
    ("activity_class", lambda k: k.activity_class or ""),
    ("is_active", lambda k: k.is_active),
    ("total_gmv_eur", lambda k: k.total_gmv),
    ("net_revenue_eur", lambda k: k.net_revenue),
    ("avg_take_rate_pct", lambda k: k.avg_take_rate),
    ("attach_rate_pct", lambda k: k.attach_rate),
    ("arpar_bg_eur", lambda k: k.arpar_bg),
    ("aov_eur", lambda k: k.aov),
    ("paid_transactions", lambda k: k.total_transactions),
    ("total_bookings", lambda k: k.total_bookings),
    ("saas_mrr_eur", lambda k: k.saas_mrr),
    ("weeks_of_data", lambda k: k.weeks_of_data),
    ("incremental_revenue_eur", lambda k: k.avg_incremental_revenue),
    ("cancellation_rate_pct", lambda k: k.avg_cancellation_rate),
]

_SUMMARY_ROWS = [
    ("active_hotels", lambda s: s.total_active_hotels),
    ("class_a_hotels", lambda s: s.class_a_hotels),
    ("class_b_hotels", lambda s: s.class_b_hotels),
    ("class_c_hotels", lambda s: s.class_c_hotels),
    ("total_rooms_live", lambda s: s.total_rooms_live),
    ("total_gmv_eur", lambda s: s.total_gmv_season),
    ("total_net_revenue_eur", lambda s: s.total_net_revenue),
    ("blended_take_rate_pct", lambda s: s.blended_take_rate),
    ("saas_mrr_eur", lambda s: s.saas_mrr),
    ("arpar_bg_eur", lambda s: s.avg_arpar_bg),
    ("attach_rate_pct", lambda s: s.portfolio_attach_rate),
    ("aov_eur", lambda s: s.avg_aov),
    ("paid_transactions", lambda s: s.total_paid_transactions),
    ("total_bookings", lambda s: s.total_bookings),
    ("incremental_revenue_per_hotel_eur", lambda s: s.avg_incremental_revenue),
]


@router.get("/csv")
def generate_csv(
    market: Optional[str] = None,
    property_type: Optional[str] = None,
    status: Optional[str] = None,
    date_from: Optional[date] = None,
    date_to: Optional[date] = None,
    db: Session = Depends(get_db),
    _: str = Depends(require_auth),
):
    """Per-hotel KPI table (full column set) plus portfolio totals."""
    hotels = _filtered_hotels(db, market, property_type, status)
    hotel_kpis = [get_hotel_kpis(db, h, date_from=date_from, date_to=date_to) for h in hotels]
    summary = get_portfolio_summary(
        db, market=market, property_type=property_type, status=status,
        date_from=date_from, date_to=date_to,
    )

    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow([name for name, _fn in _HOTEL_COLS])
    for k in hotel_kpis:
        writer.writerow([fn(k) for _name, fn in _HOTEL_COLS])

    writer.writerow([])
    writer.writerow(["--- Portfolio Totals ---"])
    for name, fn in _SUMMARY_ROWS:
        writer.writerow([name, fn(summary)])

    return Response(
        content=buf.getvalue(),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=bookinguru-export.csv"}
    )


@router.get("/xlsx")
def generate_xlsx(
    market: Optional[str] = None,
    property_type: Optional[str] = None,
    status: Optional[str] = None,
    date_from: Optional[date] = None,
    date_to: Optional[date] = None,
    db: Session = Depends(get_db),
    _: str = Depends(require_auth),
):
    """
    Excel workbook for investor modelling:
      Sheet 1 · Portfolio Summary — all top-level KPIs
      Sheet 2 · Hotels            — full KPI set per hotel
      Sheet 3 · Weekly Performance — raw weekly rows per hotel
      Sheet 4 · Daily Bookings    — raw booking-level data
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    hotels = _filtered_hotels(db, market, property_type, status)
    hotel_ids = [h.id for h in hotels]
    hmap = {h.id: h.name for h in hotels}
    hotel_kpis = [get_hotel_kpis(db, h, date_from=date_from, date_to=date_to) for h in hotels]
    summary = get_portfolio_summary(
        db, market=market, property_type=property_type, status=status,
        date_from=date_from, date_to=date_to,
    )

    wb = Workbook()
    head_font = Font(bold=True, color="FFFFFF")
    head_fill = PatternFill("solid", fgColor="0F172A")

    def style_header(ws, ncols):
        for c in range(1, ncols + 1):
            cell = ws.cell(row=1, column=c)
            cell.font = head_font
            cell.fill = head_fill
        ws.freeze_panes = "A2"

    def autosize(ws, widths):
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w

    # ── Sheet 1: Portfolio Summary ──
    ws = wb.active
    ws.title = "Portfolio Summary"
    ws.append(["Metric", "Value"])
    for name, fn in _SUMMARY_ROWS:
        ws.append([name, fn(summary)])
    ws.append(["period", _period_label(date_from, date_to)])
    ws.append(["generated", date.today().isoformat()])
    style_header(ws, 2)
    autosize(ws, [36, 22])

    # ── Sheet 2: Hotels (full KPI set) ──
    ws = wb.create_sheet("Hotels")
    ws.append([name for name, _fn in _HOTEL_COLS])
    for k in sorted(hotel_kpis, key=lambda x: -(x.total_gmv or 0)):
        ws.append([fn(k) for _name, fn in _HOTEL_COLS])
    style_header(ws, len(_HOTEL_COLS))
    autosize(ws, [28, 14, 14, 11, 14, 12, 9] + [13] * (len(_HOTEL_COLS) - 7))

    # ── Sheet 3: Weekly Performance (raw, for modelling) ──
    ws = wb.create_sheet("Weekly Performance")
    ws.append(["hotel", "week_start", "gmv_eur", "paid_transactions",
               "net_revenue_eur", "take_rate_pct", "avg_cart_value_eur"])
    if hotel_ids:
        wq = db.query(WeeklyPerformance).filter(WeeklyPerformance.hotel_id.in_(hotel_ids))
        if date_from: wq = wq.filter(WeeklyPerformance.week_start_date >= date_from)
        if date_to:   wq = wq.filter(WeeklyPerformance.week_start_date <= date_to)
        for p in wq.order_by(WeeklyPerformance.week_start_date, WeeklyPerformance.hotel_id).all():
            ws.append([
                hmap.get(p.hotel_id, "Unknown"), p.week_start_date.isoformat(),
                round(p.gmv or 0, 2), p.transactions or 0,
                round(p.bg_gross_revenue or 0, 2), round(p.take_rate or 0, 2),
                round(p.avg_cart_value or 0, 2),
            ])
    style_header(ws, 7)
    autosize(ws, [28, 12, 12, 16, 14, 13, 16])

    # ── Sheet 4: Daily Bookings (raw) ──
    ws = wb.create_sheet("Daily Bookings")
    ws.append(["date", "hotel", "vertical", "product", "vendor",
               "total_paid_eur", "service_fees_eur", "status", "channel"])
    bq = db.query(DailyBooking)
    if date_from: bq = bq.filter(DailyBooking.reservation_date >= date_from)
    if date_to:   bq = bq.filter(DailyBooking.reservation_date <= date_to)
    for b in bq.order_by(DailyBooking.reservation_date).all():
        ws.append([
            b.reservation_date.isoformat(), b.hotel_name_raw or "",
            b.vertical_name or "", b.product_name or "", b.company_name or "",
            round(b.total_paid or 0, 2), round(b.service_fees or 0, 2),
            b.reservation_status or "", b.source_channel or "",
        ])
    style_header(ws, 9)
    autosize(ws, [11, 26, 13, 30, 22, 14, 15, 12, 10])

    out = io.BytesIO()
    wb.save(out)
    return Response(
        content=out.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=bookinguru-report.xlsx"}
    )
